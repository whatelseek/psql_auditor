"""Tests for Word / Excel audit report exports."""

from pathlib import Path

from auditor.evidence_store import EvidenceStore
from auditor.report_exports import (
    parse_report_rows,
    write_docx_report,
    write_report_exports,
    write_xlsx_report,
)

SAMPLE = """# Audit Report: Demo CIS

Executive summary: two findings.

## Summary table

| ID | Title | Severity | Status | Observation | Recommendation |
|---|---|---|---|---|---|
| REQ-001 | Password encryption | High | fail | md5 in use | Use scram-sha-256 |
| REQ-002 | SSL enabled | High | pass | ssl = on | Keep TLS |
"""


def test_parse_report_rows():
    rows = parse_report_rows(SAMPLE)
    assert len(rows) == 2
    assert rows[0]["req_id"] == "REQ-001"
    assert rows[0]["status"] == "fail"
    assert "md5" in rows[0]["observation"]
    assert rows[1]["status"] == "pass"


def test_write_docx_and_xlsx(tmp_path: Path):
    docx = write_docx_report(tmp_path / "report.docx", SAMPLE)
    xlsx = write_xlsx_report(tmp_path / "report.xlsx", SAMPLE)
    assert docx.is_file() and docx.stat().st_size > 100
    assert xlsx.is_file() and xlsx.stat().st_size > 100

    from openpyxl import load_workbook

    wb = load_workbook(xlsx)
    assert "Findings" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    findings = wb["Findings"]
    assert findings["A2"].value == "REQ-001"
    assert findings["D2"].value == "fail"


def test_write_root_report_exports_sidecar_files(tmp_path: Path):
    store = EvidenceStore(tmp_path / "ClientA")
    store.write_root_report(SAMPLE)
    assert (store.root / "report.md").is_file()
    assert (store.root / "report.docx").is_file()
    assert (store.root / "report.xlsx").is_file()


def test_write_report_exports_reads_md_when_body_omitted(tmp_path: Path):
    (tmp_path / "report.md").write_text(SAMPLE, encoding="utf-8")
    out = write_report_exports(tmp_path)
    assert out["docx"] and out["xlsx"]
    assert out["error"] is None
